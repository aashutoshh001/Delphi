"""Loads the real SHL sample assessment cohort (`Book1_standardized.xlsx`:
402 candidates x 235 columns) into an `OrganizationProfile` +
`EmployeeDataLandscape` — schema-level only, per the architecture's boundary
(docs/ARCHITECTURE.md §1): the Hypothesis Agent never sees row-level employee
data, only what kinds of attributes exist and how complete each one is.

`Book1_standardized.xlsx` columns (confirmed by direct inspection — see
insight_pipeline/framework/schema.py, which classifies the same file for
the downstream Investigation Pipeline):

- `candidate_id` (dropped — not schema info)
- `{1-8}_personality` (8) — OPQ Great-8 domain scores, already aggregated
- `{domain}.{sub}_personality` (28) — OPQ facet scores feeding the domains
- `{domain}.{group}.{item}_skill` (96) — GSA self-report skill items, 8 domains
- `LC.{category}.{item}_pct` (27) — Leader-Challenge percentile fit, 4 categories
- `MQ.{facet}.{n}_raw`/`_cat` (18+18) — Motivation Questionnaire items, 4 facets
- `EL.{1-3}_sten` (3) — Enterprise-Leadership domain composites
- `HIPO.{1-2}_bucket` (2) — High-Potential composite scores
- `360.{1-8}_{rater}` (40) — 8 dimensions x 5 rater perspectives
- `tenure_years`, `compensation_annual_usd` — HR outcomes

The column-family regexes below are deliberately duplicated from (not
imported from) insight_pipeline.framework.schema: hypothesis_agent must
never import insight_pipeline (one-way dependency, enforced by
tests/unit/test_architecture_boundaries.py), so this is the
hypothesis-agent-side copy of the same real, verifiable structure — never a
fabricated competency label, same discipline as the Investigation Pipeline's
copy."""

from __future__ import annotations

import re
from pathlib import Path

from hypothesis_agent.contracts.organization import (
    AttributeField,
    EmployeeDataLandscape,
    OrganizationProfile,
)

ORGANIZATION_ID = "shl-sample-cohort"

# hypothesis_agent/src/hypothesis_agent/adapters/shl_sample_cohort.py -> Delphi/
# repo root — same anchoring pattern as config/settings.py, so this loads
# correctly regardless of the server process's launch directory.
_REPO_ROOT = Path(__file__).resolve().parents[4]
_DEFAULT_XLSX_NAME = "Book1_standardized.xlsx"

_OUTCOME_NAMES = {"tenure_years", "compensation_annual_usd"}

_PATTERNS: list[tuple[str, re.Pattern, str]] = [
    ("opq_domain", re.compile(r"^\d+_personality$"), "OPQ Great-8 domain score (personality-derived potential, already aggregated), 0-5 scale."),
    ("opq_facet", re.compile(r"^\d+\.\d+_personality$"), "OPQ facet score feeding a Great-8 domain, 0-5 scale."),
    ("gsa_skill_item", re.compile(r"^\d+\.\d+\.[a-z]_skill$"), "GSA self-report skill item (current-behaviour, stable 12-18mo), 0-5 scale."),
    ("leader_challenge_fit", re.compile(r"^LC\.\d+\.[a-z]_pct$"), "OPQ-derived percentile fit for one of SHL's Leader Challenges."),
    ("motivation_item", re.compile(r"^MQ\.[A-Z]\.\d+_(raw|cat)$"), "Motivation Questionnaire item — motivational driver, not a skill."),
    ("enterprise_leadership_domain", re.compile(r"^EL\.\d+_sten$"), "Enterprise Leadership domain composite (sten score)."),
    ("hipo_composite", re.compile(r"^HIPO\.\d+_bucket$"), "High-Potential composite score (Aspiration/Ability model)."),
    ("rater_360", re.compile(r"^360\.\d+_(self|manager|report|colleague|other)$"), "360-degree rating — a perception, not a test score."),
]


def _column_category(name: str) -> tuple[str, str, str]:
    """category, data_type, description for one column name — structural
    classification only, never a fabricated specific competency label."""
    if name in _OUTCOME_NAMES:
        return "outcome", "numeric", "HR outcome column."
    for category, pattern, description in _PATTERNS:
        if pattern.match(name):
            data_type = "categorical" if category == "motivation_item" and name.endswith("_cat") else "numeric"
            return category, data_type, description
    return "other", "unknown", "Column present in the data but not matched to a known SHL assessment family."


def load_organization_and_landscape(
    xlsx_path: Path | str | None = None,
) -> tuple[OrganizationProfile, EmployeeDataLandscape]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "load_shl_sample_cohort requires 'openpyxl': pip install openpyxl"
        ) from exc

    path = Path(xlsx_path) if xlsx_path is not None else (_REPO_ROOT / _DEFAULT_XLSX_NAME)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    row_iter = worksheet.iter_rows(values_only=True)
    header = next(row_iter)
    data_rows = list(row_iter)
    workbook.close()

    column_names = header[1:]  # drop candidate_id
    non_null_counts = [0] * len(column_names)
    total = 0
    for row in data_rows:
        total += 1
        for i, value in enumerate(row[1:]):
            if value is not None and value != "":
                non_null_counts[i] += 1

    fields: list[AttributeField] = []
    for i, name in enumerate(column_names):
        category, data_type, description = _column_category(str(name))
        coverage = non_null_counts[i] / total if total else 0.0
        fields.append(
            AttributeField(
                name=name,
                category=category,
                data_type=data_type,
                coverage_ratio=round(coverage, 4),
                description=description,
            )
        )

    landscape = EmployeeDataLandscape(
        organization_id=ORGANIZATION_ID,
        employee_count_estimate=total,
        available_fields=fields,
        notes=(
            "Sourced from an SHL assessment + HR outcomes export: GSA (current-behaviour "
            "skills), OPQ (personality-derived potential), Leader-Challenge fit, Motivation "
            "Questionnaire, Enterprise-Leadership domains, HiPo composites, 360-degree "
            "multi-rater ratings, and HR outcomes (tenure, compensation)."
        ),
    )
    profile = OrganizationProfile(
        organization_id=ORGANIZATION_ID,
        name="SHL Sample Assessment Cohort",
        core_attributes={
            "industry": "talent assessment / people analytics (sample candidate cohort)",
            "data_source": "SHL assessment export (GSA/OPQ/MQ/Leader-Challenges/HiPo/360) + HR outcomes",
            "competency_framework": "SHL Universal Competency Framework (UCF)",
            "structure": f"{total} assessed candidates; no formal org hierarchy in this sample",
            "business_goals": [
                "identify high-potential talent for leadership pipelines",
                "understand which competencies predict fit for specific business contexts",
                "understand how assessed traits relate to real HR outcomes (tenure, compensation)",
            ],
        },
    )
    return profile, landscape


if __name__ == "__main__":
    profile, landscape = load_organization_and_landscape()
    print(f"organization_id={profile.organization_id!r} name={profile.name!r}")
    print(f"employee_count_estimate={landscape.employee_count_estimate}")
    print(f"available_fields={len(landscape.available_fields)}")
    for category in sorted(landscape.categories()):
        count = sum(1 for f in landscape.available_fields if f.category == category)
        print(f"  {category}: {count} fields")
